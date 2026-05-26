#!/usr/bin/env python3
"""
Fix Excel Column Widths - Ajusta ancho de columnas en archivos Excel existentes
Uso: python fix_excel_columns.py archivo.xlsx
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def fix_excel_columns(input_file, output_file=None):
    """
    Abre un archivo Excel existente y ajusta el ancho de las columnas
    automáticamente basado en el contenido.
    
    Args:
        input_file: Ruta del archivo Excel de entrada
        output_file: Ruta del archivo Excel de salida (por defecto: input_file)
    """
    
    if output_file is None:
        output_file = input_file
    
    try:
        # Cargar workbook
        wb = openpyxl.load_workbook(input_file)
        print(f"✓ Archivo abierto: {input_file}")
        print(f"✓ Hojas encontradas: {wb.sheetnames}")
        
        # Procesar cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📋 Procesando hoja: {sheet_name}")
            
            # Calcular ancho óptimo para cada columna
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            # Calcular longitud del contenido
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Establecer ancho (aproximadamente 1.2 caracteres por unidad Excel)
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 para no exagerar
                ws.column_dimensions[column_letter].width = adjusted_width
                
                print(f"  Columna {column_letter}: ancho = {adjusted_width}")
            
            # Habilitar text wrapping en todas las celdas
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.alignment = Alignment(
                            horizontal='left',
                            vertical='top',
                            wrap_text=True
                        )
            
            # Formatear headers (primera fila) si existen
            for cell in ws[1]:
                if cell.value:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
            
            print(f"✓ Hoja {sheet_name} formateada")
        
        # Guardar archivo
        wb.save(output_file)
        print(f"\n✅ Archivo guardado: {output_file}")
        print("✓ Columnas ajustadas automáticamente")
        print("✓ Headers formateados (negrita, azul)")
        print("✓ Text wrapping habilitado")
        
    except FileNotFoundError:
        print(f"❌ Error: Archivo no encontrado: {input_file}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error al procesar archivo: {e}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python fix_excel_columns.py archivo.xlsx [salida.xlsx]")
        print("\nEjemplos:")
        print("  python fix_excel_columns.py Reporte_Distribuidora_Chavez.xlsx")
        print("  python fix_excel_columns.py entrada.xlsx salida.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    fix_excel_columns(input_file, output_file)
